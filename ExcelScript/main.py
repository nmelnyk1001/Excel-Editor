from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os

# Hard code number (typed in) – 0,0,255 (RGB number)
# Formula (calculated) – Auto Black (0,0,0)
# Linked number (pulling from another tab or sheet) – 0,153,0

# Fills to use later
hard_font = Font(color="0000FF")
formula_font = Font(color="000000")
linked_font = Font(color="009900")

# Get user directory
directory = input("Spreadsheet directory: ")
os.chdir(directory)

# Get user input file name
workbook = input("Spreadsheet name(include .xlsx): ")
wb = load_workbook(workbook)

# Create a list of sheet names
sheet_names = wb.sheetnames

# Boolean Flag for detecting references
flag = False

# Iterate through each sheet
for sheet in wb:
    # Loop through every sheet, row, and column
    for row in sheet.iter_rows():
        for cell in row:
            flag = False

            # Detect for Hard Coded Numbers
            if isinstance(cell.value, int):
                cell.font = hard_font
            elif isinstance(cell.value, float):
                cell.font = hard_font
            # Detect for References and Formulas
            elif isinstance(cell.value, str):
                if "=" in cell.value:
                    # Loop through each sheet name to see if they are referenced in the cell value
                    for name in sheet_names:
                        if name in cell.value:
                            flag = True
                            cell.font = linked_font
                    # If str contains = and does not reference another sheet, must be a formula
                    if not flag:
                        cell.font = formula_font

# Create Save Directory
save_to = directory + "\\" + workbook
wb.save(workbook)
print("Success!")

